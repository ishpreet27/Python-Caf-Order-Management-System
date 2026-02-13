import tkinter as tk
from tkinter import messagebox, simpledialog
import openpyxl
from datetime import datetime
import os

EXCEL_PATH = r"C:\Users\ISHPREET\OneDrive\Desktop\mini py\orders.xlsx"

menu = {
    "Pizza": 40,
    "Pasta": 50,
    "Burger": 60,
    "Salad": 70,
    "Coffee": 80
}

if not os.path.exists(EXCEL_PATH):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    ws.append(["Order ID", "Item", "qty", "price", "total bill", "timestamp"])
    wb.save(EXCEL_PATH)

def get_next_order_id():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    return ws.max_row

def place_order():
    order_text = order_entry.get().strip()
    if not order_text:
        messagebox.showwarning("Input Error", "Please enter your order (e.g. Pizza, Burger)")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    total_bill = 0
    order_id = get_next_order_id()

    while True:
        items = [item.strip().capitalize() for item in order_text.split(",")]

        for item in items:
            if item not in menu:
                messagebox.showwarning("Invalid Item", f"{item} is not on the menu.")
                continue
            
            qty = simpledialog.askinteger("Quantity", f"Enter quantity for {item}:", minvalue=1)
            if qty is None:
                continue
            
            price = menu[item]
            total = qty * price
            total_bill += total
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([order_id, item, qty, price, total, timestamp])
        
        more = messagebox.askyesno("Add More", "Do you want to add more items?")
        if more:
            order_text = simpledialog.askstring("Add More Items", "Enter new items (comma separated):")
            if not order_text:
                break
        else:
            break

    wb.save(EXCEL_PATH)

    if total_bill > 0:
        messagebox.showinfo("Order Placed", f"Order ID: {order_id}\nTotal Bill: Rs {total_bill}")
    order_entry.delete(0, tk.END)

def exit_app():
    root.destroy()

root = tk.Tk()
root.title("Python Café - Order System")
root.geometry("420x480")
root.config(bg="#f9f9f9")

tk.Label(root, text="Welcome to Python Café!", font=("Arial", 18, "bold"), bg="#f9f9f9").pack(pady=10)
tk.Label(root, text="Menu:", font=("Arial", 14, "bold"), bg="#f9f9f9").pack()

for item, price in menu.items():
    tk.Label(root, text=f"{item} - Rs {price}", font=("Arial", 12), bg="#f9f9f9").pack()

tk.Label(root, text="\nEnter items (comma separated):", font=("Arial", 12), bg="#f9f9f9").pack()
order_entry = tk.Entry(root, width=40, font=("Arial", 12))
order_entry.pack(pady=5)

tk.Button(root, text="Place Order", bg="green", fg="white", font=("Arial", 12, "bold"), command=place_order).pack(pady=15)
tk.Button(root, text="Exit", bg="red", fg="white", font=("Arial", 12, "bold"), command=exit_app).pack(pady=5)

root.mainloop()
